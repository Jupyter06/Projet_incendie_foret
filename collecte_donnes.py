"""
collecte_donnees.py — Collecte de données pour modèle IA
=========================================================
Corrélé avec : simulation_feu.py (classes OOP) + simulation_feu.html (GUI)

Usage :
  python collecte_donnees.py              → menu interactif
  python collecte_donnees.py --n 50       → 50 simulations rapides
  python collecte_donnees.py --n 100 --mode grille  → balayage systématique

Données collectées par simulation (features ML) :
  Entrées  — taille grille, densité p, vent, météo, mode torique
  Sorties  — étapes, arbres brûlés, pic de feu, taux de destruction…

Fichier produit : donnees_simulations.xlsx (même dossier)
  - Feuille "Simulations"  : une ligne par simulation
  - Feuille "Résumé"       : statistiques agrégées avec formules
  - Feuille "Guide ML"     : correspondance features / targets pour l'IA
"""

import os
import sys
import random
import math
from datetime import datetime
from pathlib import Path
from typing import Optional

# Import des classes OOP depuis le fichier corrélé
sys.path.insert(0, os.path.dirname(__file__))
from simulation_feu import Arbre, Vent, Meteo, Milieu, ForetTorique

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule


# ================================================================
#  PARAMÈTRES PAR DÉFAUT & PLAGES DE VARIATION
# ================================================================

PLAGES = {
    'nb_lignes':        (15, 60),
    'nb_colonnes':      (15, 60),
    'p':                (0.30, 0.95),
    'vent_vitesse':     (0, 3),
    'humidite':         (5, 95),
    'temperature':      (5, 48),
}
DIRECTIONS = ['N', 'S', 'E', 'O', 'NE', 'NO', 'SE', 'SO']

FICHIER_EXCEL = Path(__file__).parent / 'donnees_simulations.xlsx'


# ================================================================
#  COLLECTEUR — classe principale
# ================================================================

class CollecteurDonnees:
    """
    Orchestre les simulations, collecte les métriques,
    et exporte vers Excel (append si le fichier existe déjà).
    """

    def __init__(self) -> None:
        self._historique: list[dict] = []

    # ---- Exécution d'une simulation unique ----

    def lancer_simulation(self, params: dict) -> dict:
        """
        Exécute une simulation complète avec les paramètres donnés.
        Retourne un dict de métriques prêtes pour l'export.
        """
        vent  = Vent(params['vent_direction'], params['vent_vitesse'])
        meteo = Meteo(params['humidite'], params['temperature'])

        if params.get('torique', False):
            foret = ForetTorique(
                params['nb_lignes'], params['nb_colonnes'],
                params['p'], vent, meteo
            )
        else:
            foret = Milieu(
                params['nb_lignes'], params['nb_colonnes'],
                params['p'], vent, meteo
            )

        # Foyer
        if params.get('foyer_centre', False):
            foret.demarrer_foyer_centre()
        else:
            foret.demarrer_foyer_aleatoire()

        # Métriques de suivi
        sains_initial = foret.nb_sains()
        pic_feu       = foret.nb_en_feu()
        etape_pic     = 0
        courbe_feu    = [foret.nb_en_feu()]  # série temporelle
        max_etapes    = 500  # garde-fou

        # Boucle de simulation
        for _ in range(max_etapes):
            if foret.fini:
                break
            foret.step()
            feu_courant = foret.nb_en_feu()
            courbe_feu.append(feu_courant)
            if feu_courant > pic_feu:
                pic_feu   = feu_courant
                etape_pic = foret.etape

        brulees     = foret.nb_brulees()
        sains_final = foret.nb_sains()
        surface     = params['nb_lignes'] * params['nb_colonnes']
        taux_destr  = brulees / sains_initial if sains_initial > 0 else 0.0
        vitesse_moy = brulees / foret.etape   if foret.etape > 0   else 0.0

        resultat = {
            # ---- Identifiant ----
            'horodatage':       datetime.now().strftime('%Y-%m-%d %H:%M:%S'),

            # ---- Paramètres d'entrée (features ML) ----
            'nb_lignes':        params['nb_lignes'],
            'nb_colonnes':      params['nb_colonnes'],
            'surface_totale':   surface,
            'p_densite':        round(params['p'], 3),
            'vent_direction':   params['vent_direction'],
            'vent_vitesse':     params['vent_vitesse'],
            'humidite':         params['humidite'],
            'temperature':      params['temperature'],
            'torique':          int(params.get('torique', False)),
            'foyer_centre':     int(params.get('foyer_centre', False)),
            'facteur_meteo':    round(meteo.facteur_propagation(), 4),

            # ---- Résultats (targets ML) ----
            'sains_initial':    sains_initial,
            'vides_initial':    surface - sains_initial,
            'etapes_totales':   foret.etape,
            'sains_final':      sains_final,
            'brulees_final':    brulees,
            'taux_destruction': round(taux_destr, 4),
            'pic_feu':          pic_feu,
            'etape_pic_feu':    etape_pic,
            'vitesse_moy':      round(vitesse_moy, 4),
            'feu_eteint':       int(foret.fini),

            # ---- Série temporelle (JSON compact) ----
            'courbe_feu_json':  str(courbe_feu[:50]),  # 50 premiers points
        }
        self._historique.append(resultat)
        return resultat

    # ---- Génération des paramètres ----

    def _params_aleatoires(self) -> dict:
        """Tire des paramètres aléatoires dans les plages définies."""
        return {
            'nb_lignes':      random.randint(*PLAGES['nb_lignes']),
            'nb_colonnes':    random.randint(*PLAGES['nb_colonnes']),
            'p':              round(random.uniform(*PLAGES['p']), 2),
            'vent_direction': random.choice(DIRECTIONS),
            'vent_vitesse':   random.randint(*PLAGES['vent_vitesse']),
            'humidite':       round(random.uniform(*PLAGES['humidite']), 1),
            'temperature':    round(random.uniform(*PLAGES['temperature']), 1),
            'torique':        random.choice([True, False]),
            'foyer_centre':   random.choice([True, False]),
        }

    def _params_grille(self, index: int, total: int) -> dict:
        """
        Balayage systématique : varie les paramètres clés de façon uniforme.
        Utile pour couvrir l'espace de paramètres de manière contrôlée.
        """
        # Divise l'espace en tranches selon l'index
        p_values   = [0.35, 0.50, 0.65, 0.75, 0.85]
        vent_vit   = [0, 1, 2, 3]
        hum_values = [15, 35, 55, 75, 90]
        temp_vals  = [10, 20, 30, 40]

        i = index
        return {
            'nb_lignes':      30,
            'nb_colonnes':    30,
            'p':              p_values[i % len(p_values)],
            'vent_direction': DIRECTIONS[i % len(DIRECTIONS)],
            'vent_vitesse':   vent_vit[i % len(vent_vit)],
            'humidite':       hum_values[i % len(hum_values)],
            'temperature':    temp_vals[i % len(temp_vals)],
            'torique':        (i % 2 == 0),
            'foyer_centre':   (i % 3 == 0),
        }

    # ---- Lancement en batch ----

    def lancer_batch(
        self,
        n: int,
        mode: str = 'aleatoire',
        callback=None
    ) -> list[dict]:
        """
        Lance n simulations consécutives.
        mode : 'aleatoire' | 'grille'
        callback(i, n, resultat) : appelé après chaque simulation (progress)
        """
        resultats = []
        for i in range(n):
            params = (
                self._params_aleatoires() if mode == 'aleatoire'
                else self._params_grille(i, n)
            )
            r = self.lancer_simulation(params)
            resultats.append(r)
            if callback:
                callback(i + 1, n, r)
        return resultats

    # ---- Export Excel ----

    def exporter_excel(
        self,
        donnees: Optional[list[dict]] = None,
        fichier: Path = FICHIER_EXCEL
    ) -> Path:
        """
        Crée ou complète le fichier Excel.
        - Si le fichier existe : ajoute les nouvelles lignes.
        - Sinon : crée le fichier avec mise en forme complète.
        Retourne le chemin du fichier produit.
        """
        data = donnees if donnees is not None else self._historique
        if not data:
            raise ValueError("Aucune donnée à exporter.")

        if fichier.exists():
            return self._ajouter_lignes(data, fichier)
        else:
            return self._creer_fichier(data, fichier)

    # ---- Création du fichier (première fois) ----

    def _creer_fichier(self, data: list[dict], fichier: Path) -> Path:
        wb = openpyxl.Workbook()

        self._creer_feuille_simulations(wb, data)
        self._creer_feuille_resume(wb, len(data))
        self._creer_feuille_guide_ml(wb)

        wb.save(fichier)
        return fichier

    def _creer_feuille_simulations(self, wb, data: list[dict]) -> None:
        ws = wb.active
        ws.title = 'Simulations'

        # ---- Colonnes et headers ----
        colonnes = [
            # (clé_dict,               header_excel,          largeur)
            ('horodatage',             'Horodatage',               20),
            ('nb_lignes',              'Nb Lignes',                10),
            ('nb_colonnes',            'Nb Colonnes',              12),
            ('surface_totale',         'Surface totale',           14),
            ('p_densite',              'Densité p',                11),
            ('vent_direction',         'Vent Direction',           14),
            ('vent_vitesse',           'Vent Vitesse',             13),
            ('humidite',               'Humidité (%)',             13),
            ('temperature',            'Température (°C)',         16),
            ('torique',                'Torique (0/1)',            14),
            ('foyer_centre',           'Foyer Centre (0/1)',       17),
            ('facteur_meteo',          'Facteur Météo',            14),
            ('sains_initial',          'Arbres Init.',             13),
            ('vides_initial',          'Cases Vides Init.',        17),
            ('etapes_totales',         'Étapes',                   10),
            ('sains_final',            'Arbres Sains Final',       19),
            ('brulees_final',          'Arbres Brûlés',            14),
            ('taux_destruction',       'Taux Destruction',         17),
            ('pic_feu',                'Pic Feu',                  10),
            ('etape_pic_feu',          'Étape Pic Feu',            14),
            ('vitesse_moy',            'Vitesse Moy (arbres/ét.)', 24),
            ('feu_eteint',             'Feu Éteint (0/1)',         17),
            ('courbe_feu_json',        'Courbe Feu (50 pts)',      40),
        ]

        # Couleurs header — thème orange/sombre
        ORANGE_FONCE  = 'C2440C'
        ORANGE_CLAIR  = 'F97316'
        BLANC         = 'FFFFFF'
        GRIS_CLAIR    = 'F8F8F8'
        GRIS_ENTETE   = 'FFF3E0'
        VERT_CLAIR    = 'E8F5E9'
        ROUGE_CLAIR   = 'FFEBEE'

        # Groupes de colonnes
        FEATURES    = {'nb_lignes','nb_colonnes','surface_totale','p_densite',
                       'vent_direction','vent_vitesse','humidite','temperature',
                       'torique','foyer_centre','facteur_meteo'}
        TARGETS     = {'etapes_totales','sains_final','brulees_final',
                       'taux_destruction','pic_feu','etape_pic_feu',
                       'vitesse_moy','feu_eteint'}

        # ---- En-tête de groupe (ligne 1) ----
        ws.merge_cells('A1:A1')
        ws['A1'] = '🔥 Simulations de Propagation de Feu — Données pour IA'
        ws['A1'].font    = Font(bold=True, size=13, color=ORANGE_CLAIR, name='Arial')
        ws['A1'].fill    = PatternFill('solid', fgColor='1A1A2E')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 28
        # Fusionner sur toute la largeur
        ws.merge_cells(f'A1:{get_column_letter(len(colonnes))}1')

        # ---- Headers (ligne 2) ----
        for col_idx, (cle, header, largeur) in enumerate(colonnes, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)

            if cle == 'horodatage':
                bg = '37474F'
            elif cle in FEATURES:
                bg = '1565C0'   # bleu — inputs ML
            elif cle in TARGETS:
                bg = '2E7D32'   # vert — outputs ML
            else:
                bg = '37474F'

            cell.font      = Font(bold=True, color=BLANC, name='Arial', size=9)
            cell.fill      = PatternFill('solid', fgColor=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = largeur

        ws.row_dimensions[2].height = 36

        # ---- Légende des couleurs (ligne 3) ----
        ws.merge_cells(f'A3:{get_column_letter(len(colonnes))}3')
        ws['A3'] = (
            '  🔵 Colonnes bleues = Features (entrées ML)     '
            '🟢 Colonnes vertes = Targets (sorties ML à prédire)'
        )
        ws['A3'].font      = Font(italic=True, size=8, color='555555', name='Arial')
        ws['A3'].fill      = PatternFill('solid', fgColor='FFF9F0')
        ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[3].height = 18

        # ---- Données (à partir de la ligne 4) ----
        DEBUT_DONNEES = 4
        for row_idx, sim in enumerate(data, start=DEBUT_DONNEES):
            bg_row = GRIS_CLAIR if row_idx % 2 == 0 else BLANC
            for col_idx, (cle, _, _) in enumerate(colonnes, start=1):
                valeur = sim.get(cle, '')
                cell   = ws.cell(row=row_idx, column=col_idx, value=valeur)
                cell.font      = Font(name='Arial', size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill      = PatternFill('solid', fgColor=bg_row)

                # Format spécifique par type
                if cle == 'p_densite':
                    cell.number_format = '0.00'
                elif cle == 'taux_destruction':
                    cell.number_format = '0.0%'
                elif cle == 'facteur_meteo':
                    cell.number_format = '0.000'
                elif cle == 'vitesse_moy':
                    cell.number_format = '0.00'
                elif cle == 'feu_eteint':
                    # Colorer en vert si éteint, rouge sinon
                    if valeur == 1:
                        cell.fill = PatternFill('solid', fgColor='C8E6C9')
                    else:
                        cell.fill = PatternFill('solid', fgColor='FFCDD2')

        # ---- Ligne de totaux / moyennes ----
        n       = len(data)
        row_tot = DEBUT_DONNEES + n
        ws.cell(row=row_tot, column=1, value='MOYENNES').font = Font(
            bold=True, name='Arial', size=9, color='FFFFFF'
        )
        ws.cell(row=row_tot, column=1).fill = PatternFill('solid', fgColor='37474F')

        NUMERIQUES = {
            'nb_lignes', 'nb_colonnes', 'surface_totale', 'p_densite',
            'vent_vitesse', 'humidite', 'temperature', 'facteur_meteo',
            'sains_initial', 'vides_initial', 'etapes_totales', 'sains_final',
            'brulees_final', 'taux_destruction', 'pic_feu', 'etape_pic_feu',
            'vitesse_moy', 'feu_eteint'
        }
        for col_idx, (cle, _, _) in enumerate(colonnes, start=2):
            if cle in NUMERIQUES:
                col_lettre = get_column_letter(col_idx)
                formule    = f'=AVERAGE({col_lettre}{DEBUT_DONNEES}:{col_lettre}{row_tot - 1})'
                cell       = ws.cell(row=row_tot, column=col_idx, value=formule)
                cell.font  = Font(bold=True, name='Arial', size=9, color='FFFFFF')
                cell.fill  = PatternFill('solid', fgColor='37474F')
                cell.alignment = Alignment(horizontal='center')
                if cle == 'taux_destruction':
                    cell.number_format = '0.0%'

        # ---- Mise en forme finale ----
        ws.freeze_panes = 'A4'   # Gèle les 3 premières lignes (titre + headers + légende)
        ws.auto_filter.ref = f'A2:{get_column_letter(len(colonnes))}{row_tot - 1}'

        # Mise en évidence conditionnelle : taux de destruction
        taux_col  = get_column_letter(
            next(i + 1 for i, (k, _, _) in enumerate(colonnes) if k == 'taux_destruction')
        )
        ws.conditional_formatting.add(
            f'{taux_col}{DEBUT_DONNEES}:{taux_col}{row_tot - 1}',
            ColorScaleRule(
                start_type='min',  start_color='63BE7B',   # vert = faible destruction
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max',    end_color='F8696B',     # rouge = forte destruction
            )
        )

    def _creer_feuille_resume(self, wb, nb_sims: int) -> None:
        ws = wb.create_sheet('Résumé')
        ws.sheet_view.showGridLines = False

        # Titre
        ws.merge_cells('A1:F1')
        ws['A1'] = '📊 Résumé statistique des simulations'
        ws['A1'].font      = Font(bold=True, size=12, color='FFFFFF', name='Arial')
        ws['A1'].fill      = PatternFill('solid', fgColor='1A1A2E')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 26

        # Compteur
        ws['A2'] = f'Nombre de simulations : {nb_sims}'
        ws['A2'].font = Font(italic=True, size=9, color='666666', name='Arial')
        ws.row_dimensions[2].height = 18

        # Headers du tableau de stats
        headers = ['Métrique', 'Min', 'Max', 'Moyenne', 'Écart-type', 'Médiane']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=9)
            cell.fill      = PatternFill('solid', fgColor='C2440C')
            cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[3].height = 22

        # Métriques à résumer
        metriques = [
            ('Étapes totales',      'P'),   # colonne dans Simulations (approx.)
            ('Arbres brûlés',       'R'),
            ('Taux de destruction', 'S'),
            ('Pic de feu',          'T'),
            ('Facteur météo',       'M'),
        ]

        noms_metriques = [
            'Étapes totales',
            'Arbres brûlés',
            'Taux de destruction (%)',
            'Pic de feu (arbres simultanés)',
            'Facteur météo moyen',
            'Densité p',
            'Humidité (%)',
            'Température (°C)',
        ]

        cols_sim = {
            'Étapes totales':               'O',
            'Arbres brûlés':                'R',
            'Taux de destruction (%)':      'S',
            'Pic de feu (arbres simultanés)': 'T',
            'Facteur météo moyen':          'M',
            'Densité p':                    'E',
            'Humidité (%)':                 'H',
            'Température (°C)':             'I',
        }

        DEBUT = 4  # données dans Simulations démarrent ligne 4
        FIN   = DEBUT + nb_sims - 1

        COULEURS_LIGNES = ['FFFFFF', 'FFF9F0']
        for row_idx, nom in enumerate(noms_metriques, start=4):
            col_sim = cols_sim.get(nom, 'A')
            ref     = f"Simulations!{col_sim}{DEBUT}:{col_sim}{FIN}"
            bg      = COULEURS_LIGNES[row_idx % 2]

            donnees_row = [
                nom,
                f'=MIN({ref})',
                f'=MAX({ref})',
                f'=AVERAGE({ref})',
                f'=STDEV({ref})',
                f'=MEDIAN({ref})',
            ]
            for col_idx, val in enumerate(donnees_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font      = Font(name='Arial', size=9)
                cell.fill      = PatternFill('solid', fgColor=bg)
                cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left',
                                           vertical='center')
                if col_idx == 1:
                    cell.font = Font(name='Arial', size=9, bold=True)
                if nom == 'Taux de destruction (%)' and col_idx > 1:
                    cell.number_format = '0.0%'

        widths = [28, 12, 12, 14, 14, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _creer_feuille_guide_ml(self, wb) -> None:
        ws = wb.create_sheet('Guide ML')
        ws.sheet_view.showGridLines = False

        BLEU   = '1565C0'
        VERT   = '2E7D32'
        ORANGE = 'C2440C'
        BLANC  = 'FFFFFF'

        ws.merge_cells('A1:D1')
        ws['A1'] = '🤖 Guide d\'utilisation pour modèle IA'
        ws['A1'].font      = Font(bold=True, size=12, color=BLANC, name='Arial')
        ws['A1'].fill      = PatternFill('solid', fgColor='1A1A2E')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 26

        # Section Features
        ws['A3'] = 'FEATURES (colonnes d\'entrée — X)'
        ws['A3'].font = Font(bold=True, color=BLANC, name='Arial', size=10)
        ws['A3'].fill = PatternFill('solid', fgColor=BLEU)
        ws.merge_cells('A3:D3')

        features = [
            ('nb_lignes / nb_colonnes',    'Taille de la grille',
             'Influence la surface disponible à brûler'),
            ('p_densite',                  'Probabilité initiale d\'arbre',
             'Variable clé : seuil de percolation ≈ 0.59 (grille carré)'),
            ('vent_direction (encodé)',     'Direction du vent (catégorielle)',
             'Encoder en angle (0-360°) ou one-hot (8 colonnes)'),
            ('vent_vitesse',               'Force du vent (0-3)',
             'Accélère la propagation dans la direction du vent'),
            ('humidite',                   'Humidité relative (%)',
             'Freine la propagation — corrélé négativement aux brûlés'),
            ('temperature',                'Température (°C)',
             'Accélère la propagation — corrélé positivement aux brûlés'),
            ('torique',                    'Grille torique (0/1)',
             'Booléen — le feu ne s\'arrête pas aux bords'),
            ('facteur_meteo',              'Coefficient météo composé',
             'Feature dérivée utile : f_hum × f_temp'),
        ]

        row = 4
        for nom, role, note in features:
            ws.cell(row=row, column=1, value=nom).font   = Font(bold=True,  name='Arial', size=9)
            ws.cell(row=row, column=2, value=role).font  = Font(name='Arial', size=9)
            ws.cell(row=row, column=3, value=note).font  = Font(italic=True, name='Arial', size=8, color='555555')
            bg = 'EEF2FF' if row % 2 == 0 else 'F5F8FF'
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=bg)
                ws.cell(row=row, column=c).alignment = Alignment(vertical='center', wrap_text=True)
            row += 1

        # Section Targets
        row += 1
        ws.cell(row=row, column=1, value='TARGETS (colonnes de sortie — y)').font = Font(
            bold=True, color=BLANC, name='Arial', size=10
        )
        ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor=VERT)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        targets = [
            ('etapes_totales',    'Durée de la simulation',
             'Régression — combien de cycles jusqu\'à extinction'),
            ('brulees_final',     'Nombre d\'arbres brûlés',
             'Régression — grandeur absolue de la destruction'),
            ('taux_destruction',  'Proportion détruite',
             'Régression principale (0-1) — cible privilégiée pour ML'),
            ('pic_feu',           'Pic de feu simultané',
             'Régression — intensité maximale atteinte'),
            ('feu_eteint',        'Feu éteint ? (0/1)',
             'Classification — prédire si le feu se propage à toute la forêt'),
        ]

        for nom, role, note in targets:
            ws.cell(row=row, column=1, value=nom).font   = Font(bold=True,  name='Arial', size=9)
            ws.cell(row=row, column=2, value=role).font  = Font(name='Arial', size=9)
            ws.cell(row=row, column=3, value=note).font  = Font(italic=True, name='Arial', size=8, color='555555')
            bg = 'F0FFF4' if row % 2 == 0 else 'F7FFF9'
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=bg)
                ws.cell(row=row, column=c).alignment = Alignment(vertical='center', wrap_text=True)
            row += 1

        # Remarque percolation
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row=row, column=1,
                value='⚠ Note : le seuil de percolation (p ≈ 0.59 pour grille carré) '
                      'est la valeur critique où la propagation devient globale. '
                      'C\'est la non-linéarité centrale du système — le modèle ML devra la capturer.').font = Font(
            italic=True, size=8, color='774400', name='Arial'
        )
        ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='FFF3E0')
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[row].height = 36

        col_widths = [28, 30, 55, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Ajout de lignes à un fichier existant ----

    def _ajouter_lignes(self, data: list[dict], fichier: Path) -> Path:
        wb = openpyxl.load_workbook(fichier)
        ws = wb['Simulations']

        # Trouver la dernière ligne de données (avant la ligne MOYENNES)
        derniere_ligne = ws.max_row
        # Chercher la ligne 'MOYENNES' et supprimer l'ancienne
        ligne_moy = None
        for row in ws.iter_rows(min_row=4):
            if row[0].value == 'MOYENNES':
                ligne_moy = row[0].row
                break
        if ligne_moy:
            ws.delete_rows(ligne_moy)

        # Nombre total de colonnes
        nb_cols = ws.max_column
        colonnes_keys = [
            'horodatage', 'nb_lignes', 'nb_colonnes', 'surface_totale',
            'p_densite', 'vent_direction', 'vent_vitesse', 'humidite',
            'temperature', 'torique', 'foyer_centre', 'facteur_meteo',
            'sains_initial', 'vides_initial', 'etapes_totales', 'sains_final',
            'brulees_final', 'taux_destruction', 'pic_feu', 'etape_pic_feu',
            'vitesse_moy', 'feu_eteint', 'courbe_feu_json',
        ]

        BLANC      = 'FFFFFF'
        GRIS_CLAIR = 'F8F8F8'

        premiere_nouvelle = ws.max_row + 1
        for row_idx_offset, sim in enumerate(data):
            row_idx = premiere_nouvelle + row_idx_offset
            bg_row  = GRIS_CLAIR if row_idx % 2 == 0 else BLANC
            for col_idx, cle in enumerate(colonnes_keys, start=1):
                cell       = ws.cell(row=row_idx, column=col_idx, value=sim.get(cle, ''))
                cell.font  = Font(name='Arial', size=9)
                cell.fill  = PatternFill('solid', fgColor=bg_row)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cle == 'taux_destruction':
                    cell.number_format = '0.0%'

        # Réajouter la ligne MOYENNES
        row_tot = ws.max_row + 1
        debut   = 4
        fin     = row_tot - 1
        ws.cell(row=row_tot, column=1, value='MOYENNES').font = Font(
            bold=True, name='Arial', size=9, color='FFFFFF'
        )
        ws.cell(row=row_tot, column=1).fill = PatternFill('solid', fgColor='37474F')
        NUMERIQUES_IDX = {2,3,4,5,7,8,9,13,14,15,16,17,18,19,20,21,22}
        for col_idx in NUMERIQUES_IDX:
            col_lettre = get_column_letter(col_idx)
            cell       = ws.cell(row=row_tot, column=col_idx,
                                 value=f'=AVERAGE({col_lettre}{debut}:{col_lettre}{fin})')
            cell.font  = Font(bold=True, name='Arial', size=9, color='FFFFFF')
            cell.fill  = PatternFill('solid', fgColor='37474F')
            cell.alignment = Alignment(horizontal='center')
            if col_idx == 18:
                cell.number_format = '0.0%'

        # Màj résumé
        if 'Résumé' in wb.sheetnames:
            ws_res = wb['Résumé']
            nb_total = fin - debut + 1
            ws_res['A2'] = f'Nombre de simulations : {nb_total}'

        wb.save(fichier)
        return fichier


# ================================================================
#  POINT D'ENTRÉE
# ================================================================

def _barre_progression(i, n, r):
    pct    = i / n * 100
    barres = int(pct / 2)
    barre  = '█' * barres + '░' * (50 - barres)
    taux   = r['taux_destruction'] * 100
    print(f'\r  [{barre}] {pct:5.1f}% — sim {i:>4}/{n} | '
          f'brûlés: {r["brulees_final"]:>5} | taux: {taux:5.1f}%', end='', flush=True)


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Collecte de données — simulations feu')
    parser.add_argument('--n',    type=int, default=None,   help='Nombre de simulations')
    parser.add_argument('--mode', type=str, default=None,   choices=['aleatoire','grille'],
                        help='Mode de variation des paramètres')
    args = parser.parse_args()

    print('\n' + '=' * 60)
    print('  Collecte de données — Propagation de Feu')
    print('  Corrélé avec : simulation_feu.py + simulation_feu.html')
    print('=' * 60)

    if FICHIER_EXCEL.exists():
        print(f'\n  📂 Fichier existant détecté : {FICHIER_EXCEL.name}')
        print('     Les nouvelles simulations seront ajoutées.')
    else:
        print(f'\n  📄 Nouveau fichier : {FICHIER_EXCEL.name}')

    # Nombre de simulations
    if args.n:
        n = args.n
    else:
        try:
            n = int(input('\n  Nombre de simulations à lancer (défaut=20) : ') or 20)
        except ValueError:
            n = 20

    # Mode
    if args.mode:
        mode = args.mode
    else:
        print('\n  Mode de variation des paramètres :')
        print('    [1] Aléatoire (Monte Carlo) — recommandé pour ML')
        print('    [2] Grille (balayage systématique)')
        choix = input('  Choix (1/2, défaut=1) : ').strip()
        mode = 'grille' if choix == '2' else 'aleatoire'

    print(f'\n  ▶ Lancement de {n} simulations en mode "{mode}"…\n')

    collecteur = CollecteurDonnees()
    t0 = datetime.now()
    resultats  = collecteur.lancer_batch(n, mode=mode, callback=_barre_progression)
    duree      = (datetime.now() - t0).total_seconds()

    print(f'\n\n  ✅ {n} simulations terminées en {duree:.1f}s')
    print(f'     Taux moy. destruction : {sum(r["taux_destruction"] for r in resultats)/n:.1%}')
    print(f'     Étapes moy.           : {sum(r["etapes_totales"] for r in resultats)/n:.1f}')
    print(f'     Feux éteints          : {sum(r["feu_eteint"] for r in resultats)}/{n}')

    print(f'\n  💾 Export Excel en cours…')
    chemin = collecteur.exporter_excel(resultats)
    print(f'  ✅ Fichier sauvegardé : {chemin.name}')
    print()


if __name__ == '__main__':
    main()